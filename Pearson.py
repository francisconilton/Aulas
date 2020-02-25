# -*- coding: utf-8 -*-
import os
from openpyxl.reader.excel import load_workbook
import re
from datetime import date
from tkinter import *
import tkinter as tk
from tkinter import filedialog as fd
from os.path import expanduser
home = expanduser("~")


def sheet(path, novo):
    patt = re.compile('\d\d\d\d-\d\d-\d\d')
    today = str(date.today())
    wb = load_workbook(path)
    ws_base = wb['Agendamento_Março']
    dic = {}
    cell_name = ws_base.cell(column=1, row=4).value
    row = 4
    while bool(cell_name):
        cell_name = ws_base.cell(column=1, row=row).value
        cell_date = ws_base.cell(column=2, row=row).value
        cell_city = ws_base.cell(column=4, row=row).value
        if cell_name in dic.keys():
            dic[cell_name][cell_date] = cell_city
        else:
            dic[cell_name] = {cell_date: cell_city}
        row +=1
    ws_marco = wb['Planilha1']
    row = 3
    cell = True
    while bool(cell):
        data = ws_marco.cell(column=1, row=row).value
        nome = True
        col = 3
        while bool(nome):
            nome = ws_marco.cell(column=col, row=2).value
            if nome in dic.keys():
                if data in dic[nome].keys():
                    r = dic[nome][data]
                    ws_marco.cell(row=row, column=col).value = r
            col += 1
            if col > 1000:break
        row += 1
        if row > 1000:break
    nomeatual = os.path.basename(path)
    wb.save(path.replace(nomeatual, novo))


master = tk.Tk()
master.geometry("+100+100")
master.title("PEARSON: Atualização mensal")
tk.Label(master, text="Arquivo base").grid(row=0)
tk.Label(master, text="Novo nome").grid(row=1)

def callback():
    name= fd.askopenfilename(initialdir=home)
    e1.delete(0, END)  # deletes the current value
    e1.insert(0, name)

def update():
    novo_nome = e2.get()
    path = e1.get()
    sheet(path, novo_nome + '.xlsx')
    master.quit()

e1 = tk.Entry(master, width=50)
e2 = tk.Entry(master, width=50)
b1 = tk.Button(text='Buscar ...', command=callback)
b2 = tk.Button(text='Atualizar', command=update)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
b1.grid(row=0, column=2)
b2.grid(row=3, column=1)

master.mainloop()




